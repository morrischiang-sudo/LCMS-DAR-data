"""
DAR Compass - ADC DAR Analysis Platform
-----------------------------------------
Streamlit front-end for dar_calculator.py.

Run with:
    streamlit run app.py
"""

from __future__ import annotations

import io
import os

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
import streamlit as st

from dar_calculator import (
    MassVariant,
    PayloadDef,
    base_masses_from_mab,
    build_drug_load_summary_table,
    build_theoretical_table,
    build_verification_table,
    calculate_dar,
    consolidate_by_total_count,
    estimate_theoretical_grid_size,
    filter_by_fractional_abundance,
    load_deconvolution_file,
    marginal_distribution_by_chemistry,
    match_species,
)

APP_TITLE = "DAR Compass"
APP_TAGLINE = "Automated DAR distribution analysis for antibody-drug conjugates"

st.set_page_config(page_title=APP_TITLE, page_icon="\U0001F9ED", layout="wide")


# --------------------------------------------------------------------------
# Access gate
# --------------------------------------------------------------------------
# Enforced only if an APP_PASSWORD is configured (e.g. as a Render
# environment variable / secret). Local development with no APP_PASSWORD
# set runs ungated, so this never gets in the way of local iteration -
# it only matters once the app is deployed somewhere with a public URL.

def _get_app_password() -> str | None:
    pw = os.environ.get("APP_PASSWORD")
    if pw:
        return pw
    try:
        return st.secrets["APP_PASSWORD"]
    except Exception:
        return None


def _require_password() -> None:
    expected = _get_app_password()
    if not expected:
        return  # no password configured - nothing to gate

    if st.session_state.get("_authenticated"):
        return

    st.title(f"\U0001F9ED {APP_TITLE}")
    st.caption(APP_TAGLINE)
    st.info("This app is password-protected. Ask whoever shared the link with you for the password.")
    pw = st.text_input("Password", type="password", key="_password_input")
    if st.button("Enter", type="primary"):
        if pw == expected:
            st.session_state["_authenticated"] = True
            st.rerun()
        else:
            st.error("Incorrect password.")
    st.stop()


_require_password()


# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------

def read_uploaded_excel(uploaded_file) -> pd.DataFrame:
    df = pd.read_excel(uploaded_file)
    required = {"Average Mass", "Sum Intensity"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"'{uploaded_file.name}' is missing expected column(s): {missing}")
    return df


def make_distribution_chart(matched_df: pd.DataFrame, dar: dict, payload_labels: list[str], top_n: int = 15):
    fig, ax = plt.subplots(figsize=(9, 4.5))
    if matched_df.empty:
        ax.text(0.5, 0.5, "No species matched at this tolerance", ha="center", va="center")
        ax.axis("off")
        return fig

    top = matched_df.sort_values("relative_abundance", ascending=False).head(top_n)
    colors = ["#C55A11" if a else "#4472C4" for a in top["ambiguous"]]
    ax.bar(top["species"], top["relative_abundance"] * 100, color=colors)
    ax.set_ylabel("Relative abundance (%)")
    ax.set_xlabel("Assigned species")
    dar_text = "  |  ".join(f"DAR[{lbl}] = {dar[lbl]:.2f}" for lbl in payload_labels)
    ax.set_title(f"{dar_text}  |  Total DAR = {dar['total']:.2f}")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    return fig


def make_drug_load_chart(long_df: pd.DataFrame, payload_defs: list[PayloadDef], dar: dict):
    n_chem = max(len(payload_defs), 1)
    fig, axes = plt.subplots(1, n_chem, figsize=(5.5 * n_chem, 4.5))
    if n_chem == 1:
        axes = [axes]
    for ax, p in zip(axes, payload_defs):
        sub = long_df[long_df["chemistry"] == p.label].sort_values("count")
        counts = sub["count"].tolist()
        pcts = sub["relative_abundance_pct"].tolist()
        colors = ["#2A8C82"] * len(pcts)
        if pcts:
            colors[pcts.index(max(pcts))] = "#1B2A4A"  # highlight the mode, matching the reference report style
        ax.bar([str(c) for c in counts], pcts, color=colors)
        ax.set_xlabel(f"{p.label} count")
        ax.set_ylabel("Relative abundance (%)")
        ax.set_title(f"{p.label}\nAverage = {dar.get(p.label, float('nan')):.2f}")
    plt.tight_layout()
    return fig


def style_drug_load_table(table: pd.DataFrame):
    count_cols = [c for c in table.columns if c != "Average"]

    def highlight_mode(row):
        styles = [""] * len(row)
        vals = row[count_cols].dropna()
        if len(vals):
            mode_col = vals.idxmax()
            styles[list(row.index).index(mode_col)] = "background-color: #CFE2F3; font-weight: bold; color: #1B2A4A;"
        return styles

    return table.style.apply(highlight_mode, axis=1).format("{:.2f}", na_rep="")


def build_params_table(
    payload_defs: list[PayloadDef],
    base_masses: list[float],
    base_mass_mode: str,
    ppm_tolerance: float,
    adc_min_fractional_abundance: float,
) -> pd.DataFrame:
    """One row per analysis parameter, in the same spirit as the header
    block of a manual analysis sheet - so someone reviewing the output can
    see exactly what was configured without having to ask.
    """
    rows = [
        {"Parameter": "mAb base mass mode", "Value": base_mass_mode},
        {"Parameter": "mAb base mass(es) used (Da)", "Value": ", ".join(f"{b:.4f}" for b in base_masses)},
        {"Parameter": "Mass accuracy tolerance (ppm)", "Value": ppm_tolerance},
        {"Parameter": "ADC fractional abundance threshold (%)", "Value": adc_min_fractional_abundance},
    ]
    for p in payload_defs:
        rows.append({"Parameter": f"Chemistry \"{p.label}\" - max conjugation count", "Value": max(p.n_values) if p.n_values else 0})
        rows.append({"Parameter": f"Chemistry \"{p.label}\" - count step", "Value": (p.n_values[1] - p.n_values[0]) if len(p.n_values) > 1 else 1})
        for v in p.variants:
            rows.append({
                "Parameter": f"Chemistry \"{p.label}\" - mass variant \"{v.variant_label}\"",
                "Value": f"MW = {v.mw:.4f} Da, DAR weight = {v.dar_weight}",
            })
    return pd.DataFrame(rows)


def write_labeled_block(ws, title: str, df: pd.DataFrame, start_row: int) -> int:
    """Write a titled table starting at `start_row`; return the next free row."""
    from openpyxl.styles import Font
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12)
    start_row += 1
    if df is None or df.empty:
        ws.cell(row=start_row, column=1, value="(none)")
        return start_row + 2

    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=j, value=str(col)).font = Font(bold=True)
    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for j, val in enumerate(row.tolist(), start=1):
            ws.cell(row=i, column=j, value=val)
    return start_row + len(df) + 3


def to_excel_bytes(
    summary_df: pd.DataFrame,
    species_df: pd.DataFrame,
    drug_load_df: pd.DataFrame | None = None,
    params_df: pd.DataFrame | None = None,
    theoretical_df: pd.DataFrame | None = None,
    verification_df: pd.DataFrame | None = None,
) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="DAR Summary", index=False)
        if drug_load_df is not None and not drug_load_df.empty:
            drug_load_df.to_excel(writer, sheet_name="Drug-load Distribution")
        species_df.to_excel(writer, sheet_name="Matched Species", index=False)

        if params_df is not None:
            ws = writer.book.create_sheet("Analysis")
            row = write_labeled_block(ws, "Analysis Parameters", params_df, 1)
            row = write_labeled_block(ws, "Theoretical Mass Combinations (all permutations)", theoretical_df, row)
            write_labeled_block(
                ws,
                "Peak-by-Peak Verification (every candidate peak, matched or not)",
                verification_df,
                row,
            )
    return buf.getvalue()


# --------------------------------------------------------------------------
# Sidebar - inputs
# --------------------------------------------------------------------------

st.sidebar.title(f"\U0001F9EA {APP_TITLE}")
st.sidebar.caption(APP_TAGLINE)

st.sidebar.header("1. Upload deconvolution files")
mab_file = st.sidebar.file_uploader(
    "Naked mAb reference export (.xlsx)",
    type=["xlsx"],
    help="Deconvoluted intact-mass export for the unconjugated antibody. Used to determine base masses (glycoform/adduct variants).",
)
adc_file = st.sidebar.file_uploader(
    "ADC export (.xlsx)",
    type=["xlsx"],
    help="Deconvoluted intact-mass export for the conjugated ADC sample.",
)

st.sidebar.header("2. Linker-payload chemistries")
n_payloads = st.sidebar.number_input("Number of distinct linker-payload chemistries", min_value=1, max_value=4, value=2, step=1)

payload_defs: list[PayloadDef] = []
for i in range(int(n_payloads)):
    with st.sidebar.expander(f"Chemistry {i + 1}", expanded=(i < 2)):
        include = st.checkbox("Include in analysis", value=True, key=f"include_{i}")
        label = st.text_input("Label", value=str(i + 1), key=f"label_{i}", help="Short tag, e.g. '4', '12', 'DXd'.")

        n_variants = st.number_input(
            "Number of mass variants for this chemistry",
            min_value=1, max_value=4, value=1, step=1, key=f"nvariants_{i}",
            help=(
                "Usually 1 (the intact linker-payload mass). Add more if breakage during harsh sample "
                "processing produces additional possible masses at a conjugation site — DAR Compass will "
                "consider any mix of these variants across a chemistry's attachment sites within one molecule."
            ),
        )

        variants: list[MassVariant] = []
        for v in range(int(n_variants)):
            variant_tag = "intact" if v == 0 else f"variant {v + 1}"
            st.markdown(f"**Mass variant {v + 1}** ({variant_tag})" if v == 0 else f"**Mass variant {v + 1}**")
            vmw = st.number_input(
                "MW (Da)", min_value=0.0, value=1717.83 if v == 0 else 0.0, step=0.01, format="%.2f",
                key=f"mw_{i}_{v}",
            )
            vweight = st.number_input(
                "DAR weight", min_value=0.0, max_value=1.0, value=1.0, step=0.05, format="%.2f",
                key=f"weight_{i}_{v}",
                help=(
                    "How much one occupied site of this mass variant counts toward DAR. 1.0 (default) = "
                    "counts as a full payload, same as an intact site. Set lower (or 0) for a variant "
                    "that represents partial or complete payload loss."
                ),
            )
            variants.append(MassVariant(mw=float(vmw), dar_weight=float(vweight)))

        max_n = st.number_input("Max conjugation count", min_value=0, max_value=20, value=8, step=1, key=f"maxn_{i}")
        step = st.selectbox("Allowed count step", options=[1, 2], index=0, key=f"step_{i}", help="Use 2 if conjugation only occurs in pairs (e.g. per disulfide bond).")
        if include:
            payload_defs.append(PayloadDef(
                label=label.strip() or str(i + 1),
                variants=variants,
                n_values=list(range(0, int(max_n) + 1, int(step))),
            ))

st.sidebar.header("3. ADC candidate selection")
adc_min_fractional_abundance = st.sidebar.number_input(
    "ADC fractional abundance threshold (%)",
    min_value=0.0, max_value=100.0, value=0.0, step=0.01, format="%.2f",
    help=(
        "Only ADC peaks with Fractional Abundance at or above this value are treated as "
        "candidate species and passed on to matching/DAR calculation. Raise this to drop "
        "very low-abundance peaks (noise, minor adducts) before matching against the "
        "theoretical mass ladder. Leave at 0 to consider every peak in the file."
    ),
)

st.sidebar.header("4. Matching parameters")
ppm_tolerance = st.sidebar.slider(
    "Mass accuracy tolerance (ppm)",
    min_value=10, max_value=500, value=300, step=10,
    help="Peaks are matched to the nearest theoretical species and kept only if within this tolerance. Validated against historical runs at ~250-350 ppm for ~150-165 kDa intact species; tune per instrument/method.",
)
base_mass_mode = st.sidebar.radio(
    "mAb base mass(es) to use",
    options=["Most abundant only (recommended)", "Top N glycoform/adduct variants"],
)
base_mass_top_n = 1
if base_mass_mode.startswith("Top N"):
    base_mass_top_n = st.sidebar.number_input("N", min_value=1, max_value=10, value=3, step=1)

run_clicked = st.sidebar.button("Run DAR analysis", type="primary", use_container_width=True)


# --------------------------------------------------------------------------
# Main panel
# --------------------------------------------------------------------------
# Results are computed only when "Run DAR analysis" is clicked, then stashed
# in st.session_state and rendered from there on every subsequent script
# run. Without this, results would disappear the moment you touched any
# other widget (like the chart-detail toggle below) - st.button() only
# reads True on the exact run it was clicked on; every other widget
# interaction reruns the script with that button reading False again.

st.title(f"\U0001F9EA {APP_TITLE}")
st.caption(APP_TAGLINE)


def run_analysis():
    if mab_file is None or adc_file is None:
        st.error("Please upload both a mAb reference file and an ADC file before running the analysis.")
        return None

    if not payload_defs:
        st.error("Please include at least one linker-payload chemistry.")
        return None

    try:
        mab_df = read_uploaded_excel(mab_file)
        adc_df_all = read_uploaded_excel(adc_file)
    except ValueError as e:
        st.error(str(e))
        return None

    try:
        adc_df = filter_by_fractional_abundance(adc_df_all, adc_min_fractional_abundance)
    except ValueError as e:
        st.error(str(e))
        return None

    if adc_df.empty:
        st.error(
            f"No ADC peaks remain at a {adc_min_fractional_abundance:.2f}% fractional abundance threshold "
            f"(out of {len(adc_df_all)} peaks in the file). Lower the threshold and try again."
        )
        return None

    base_masses = base_masses_from_mab(mab_df, top_n=int(base_mass_top_n))
    estimated_grid_size = estimate_theoretical_grid_size(base_masses, payload_defs)
    theoretical = build_theoretical_table(base_masses, payload_defs)
    matched = match_species(adc_df, theoretical, ppm_tolerance=ppm_tolerance)
    dar, matched_with_contrib = calculate_dar(matched, payload_defs)

    # Column name for every configured mass variant (n_<variant_label>), plus a
    # "_total" column for chemistries with more than one variant.
    n_display_cols: list[str] = []
    for p in payload_defs:
        for v in p.variants:
            n_display_cols.append(f"n_{v.variant_label}")
        if len(p.variants) > 1:
            n_display_cols.append(f"n_{p.label}_total")

    verification_df = build_verification_table(
        adc_df_all, theoretical, ppm_tolerance,
        abundance_threshold=adc_min_fractional_abundance,
    )

    return {
        "payload_defs": payload_defs,
        "payload_labels": [p.label for p in payload_defs],
        "dar": dar,
        "matched_with_contrib": matched_with_contrib,
        "base_masses": base_masses,
        "base_mass_mode": base_mass_mode,
        "theoretical": theoretical,
        "verification_df": verification_df,
        "n_observed": len(adc_df_all),
        "n_candidates": len(adc_df),
        "n_matched": len(matched_with_contrib),
        "n_display_cols": n_display_cols,
        "adc_min_fractional_abundance": adc_min_fractional_abundance,
        "ppm_tolerance": ppm_tolerance,
        "estimated_grid_size": estimated_grid_size,
    }


if run_clicked:
    result = run_analysis()
    if result is not None:
        st.session_state["dar_compass_results"] = result

if "dar_compass_results" not in st.session_state:
    st.info(
        "Upload a mAb reference file and an ADC file in the sidebar, confirm your linker-payload "
        "chemistries and tolerance, then click **Run DAR analysis**.\n\n"
        "Expected file format: a deconvoluted intact-mass export (e.g. Thermo BioPharma Finder) "
        "with at least `Average Mass` and `Sum Intensity` columns."
    )
    st.stop()

r = st.session_state["dar_compass_results"]
payload_defs = r["payload_defs"]
payload_labels = r["payload_labels"]
dar = r["dar"]
matched_with_contrib = r["matched_with_contrib"]
base_masses = r["base_masses"]
base_mass_mode = r["base_mass_mode"]
theoretical = r["theoretical"]
verification_df = r["verification_df"]
n_observed = r["n_observed"]
n_candidates = r["n_candidates"]
n_matched = r["n_matched"]
n_display_cols = r["n_display_cols"]
adc_min_fractional_abundance = r["adc_min_fractional_abundance"]
ppm_tolerance = r["ppm_tolerance"]
estimated_grid_size = r["estimated_grid_size"]

if estimated_grid_size > 300_000:
    st.warning(
        f"This configuration builds roughly {estimated_grid_size:,} theoretical species, which may run "
        "slowly. Consider reducing the number of mass variants, max conjugation count, or mAb base masses."
    )

with st.expander("Mass variants configured for this run"):
    variant_rows = []
    for p in payload_defs:
        for v in p.variants:
            variant_rows.append({
                "Chemistry": p.label, "Variant": v.variant_label,
                "MW (Da)": v.mw, "DAR weight": v.dar_weight,
            })
    st.dataframe(pd.DataFrame(variant_rows), use_container_width=True, hide_index=True)

with st.expander("Peak-by-peak verification (every peak in the ADC file, matched or not)"):
    st.caption(
        "Every peak from the uploaded ADC file, with its closest theoretical species regardless of "
        "whether it was actually accepted - so you can check *why* a peak wasn't counted (excluded by "
        "the abundance threshold vs. simply too far in ppm from anything in the theoretical grid) "
        "instead of it just silently disappearing. Green = matched. Yellow = excluded by the "
        "fractional abundance threshold before matching was even attempted. Red = passed the "
        "abundance threshold but its closest theoretical species was still outside the ppm tolerance "
        "- if this happens for a peak you expected to match, it usually means a chemistry's max "
        "conjugation count (or a mass variant's MW) needs adjusting, not that the peak is real noise."
    )

    def highlight_verification(row):
        if row["matched"]:
            color = "#D9EAD3"  # green
        elif not row["passed_abundance_threshold"]:
            color = "#FFF2CC"  # yellow
        else:
            color = "#F4CCCC"  # red
        return [f"background-color: {color}" for _ in row]

    verification_display_cols = [
        "Average Mass", "Sum Intensity", "Fractional Abundance", "closest_species",
        "closest_theoretical_mass", "delta_mass", "ppm_error",
        "passed_abundance_threshold", "within_ppm_tolerance", "matched",
    ]
    verification_display_cols = [c for c in verification_display_cols if c in verification_df.columns]
    st.dataframe(
        verification_df[verification_display_cols].style.apply(highlight_verification, axis=1),
        use_container_width=True,
    )

# --- Top-line metrics -------------------------------------------------
cols = st.columns(len(payload_labels) + 1)
for col, lbl in zip(cols, payload_labels):
    col.metric(f"DAR [{lbl}]", f"{dar[lbl]:.2f}")
cols[-1].metric("Total DAR", f"{dar['total']:.2f}")

n_ambiguous = int(matched_with_contrib["ambiguous"].sum()) if n_matched else 0

filter_note = (
    f" ({n_observed} peaks in file → {n_candidates} at or above {adc_min_fractional_abundance:.2f}% fractional abundance)"
    if adc_min_fractional_abundance > 0
    else f" ({n_observed} peaks in file, no fractional abundance filter applied)"
)
st.caption(
    f"Matched {n_matched} of {n_candidates} candidate ADC peaks to an intact theoretical species "
    f"within {ppm_tolerance} ppm{filter_note}. {n_ambiguous} matched peak(s) flagged as ambiguous."
)

if n_ambiguous:
    st.warning(
        f"{n_ambiguous} matched peak(s) had more than one plausible species within tolerance. "
        "Review the highlighted rows below before trusting the DAR value."
    )

# --- Drug-load distribution (per-chemistry summary table + chart) --------
st.subheader("Drug-load distribution")
st.caption(
    "Percent of matched intensity at each individual count, per chemistry - independent of the other "
    "chemistries' counts. This is the standard drug-load-distribution report format (one row per "
    "linker-payload, one column per count, plus an average)."
)

drug_load_long = marginal_distribution_by_chemistry(matched_with_contrib, payload_defs)
drug_load_table = build_drug_load_summary_table(matched_with_contrib, payload_defs, dar)

has_nondefault_dar_weight = any(v.dar_weight != 1.0 for p in payload_defs for v in p.variants)
if has_nondefault_dar_weight:
    st.caption(
        "Note: at least one mass variant has a DAR weight other than 1.0. \"Average\" is each "
        "chemistry's actual DAR (weighted by dar_weight, same number as the metric above) - it will "
        "differ from the simple mean of the row above it, which counts every occupied site equally "
        "regardless of variant."
    )

if not drug_load_table.empty:
    st.dataframe(style_drug_load_table(drug_load_table), use_container_width=True)
    drug_load_fig = make_drug_load_chart(drug_load_long, payload_defs, dar)
    st.pyplot(drug_load_fig)
else:
    drug_load_fig = None
    st.write("No matched species to summarize yet.")

# --- Chart --------------------------------------------------------------
st.subheader("DAR distribution (species-level)")

has_multi_variant_chemistry = any(len(p.variants) > 1 for p in payload_defs)
chart_df = matched_with_contrib
if has_multi_variant_chemistry:
    chart_view = st.radio(
        "Chart detail",
        options=["Detailed (show mass variants)", "Consolidated (total counts per chemistry)"],
        horizontal=True,
        help=(
            "Detailed shows every intact/broken mixture as its own bar - precise, but can get "
            "cluttered once a chemistry has multiple mass variants. Consolidated groups bars by "
            "total occupied-site count per chemistry only, hiding which specific variant(s) made "
            "up that total. This toggle only changes this chart - the table below and the DAR "
            "numbers above are unaffected either way."
        ),
    )
    if chart_view.startswith("Consolidated"):
        chart_df = consolidate_by_total_count(matched_with_contrib, payload_defs)

fig = make_distribution_chart(chart_df, dar, payload_labels)
st.pyplot(fig)

# --- Matched species table ----------------------------------------------
st.subheader("Matched species")
if n_matched:
    display_cols = ["Average Mass", "Sum Intensity", "species", "ppm_error", "relative_abundance"]
    display_cols += n_display_cols
    display_cols += ["ambiguous", "runner_up_species", "runner_up_ppm_error"]
    display_df = matched_with_contrib[display_cols].copy()
    display_df["relative_abundance"] = (display_df["relative_abundance"] * 100).round(2)
    display_df = display_df.rename(columns={"relative_abundance": "relative_abundance_%"})

    def highlight_ambiguous(row):
        return ["background-color: #FCE4D6" if row["ambiguous"] else "" for _ in row]

    st.dataframe(display_df.style.apply(highlight_ambiguous, axis=1), use_container_width=True)
else:
    st.write("No peaks matched at this tolerance. Try increasing the ppm tolerance.")

# --- Downloads ------------------------------------------------------------
st.subheader("Download results")
summary_df = pd.DataFrame(
    {
        "metric": [f"DAR [{lbl}]" for lbl in payload_labels] + [
            "Total DAR", "Matched peaks", "Candidate peaks (after abundance filter)",
            "Observed peaks (raw file)", "ADC fractional abundance threshold (%)",
            "Ambiguous matches", "ppm tolerance used",
        ],
        "value": [round(dar[lbl], 3) for lbl in payload_labels] + [
            round(dar["total"], 3), n_matched, n_candidates,
            n_observed, adc_min_fractional_abundance,
            n_ambiguous, ppm_tolerance,
        ],
    }
)

params_df = build_params_table(payload_defs, base_masses, base_mass_mode, ppm_tolerance, adc_min_fractional_abundance)

col1, col2, col3 = st.columns(3)
with col1:
    st.download_button(
        "Download DAR report (.xlsx)",
        data=to_excel_bytes(
            summary_df, matched_with_contrib, drug_load_table,
            params_df=params_df, theoretical_df=theoretical, verification_df=verification_df,
        ),
        file_name="dar_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        help=(
            "Includes the drug-load distribution table and an \"Analysis\" sheet with every parameter "
            "used, the full theoretical mass grid, and a peak-by-peak verification table - so results "
            "can be checked by hand."
        ),
    )
with col2:
    img_buf = io.BytesIO()
    fig.savefig(img_buf, format="png", dpi=150)
    st.download_button(
        "Download species-level chart (.png)",
        data=img_buf.getvalue(),
        file_name="dar_distribution_species.png",
        mime="image/png",
        use_container_width=True,
    )
with col3:
    if drug_load_fig is not None:
        drug_load_img_buf = io.BytesIO()
        drug_load_fig.savefig(drug_load_img_buf, format="png", dpi=150)
        st.download_button(
            "Download drug-load chart (.png)",
            data=drug_load_img_buf.getvalue(),
            file_name="drug_load_distribution.png",
            mime="image/png",
            use_container_width=True,
        )

with st.expander("Base masses used for this run"):
    st.write(pd.DataFrame({"base_mass": base_masses}))
