"""
dar_ui_helpers.py
------------------
Shared Streamlit UI building blocks for DAR Compass's two analysis pages
(SEC LC-MS Analysis and RP LC-MS / IdeZ digestion Analysis). Kept separate
from dar_calculator.py so that module stays pure Python/pandas with no
Streamlit dependency (easy to unit-test and reuse outside the app).

Everything here is either:
  - the password gate (shared across both pages),
  - a chart/table rendering helper, or
  - `run_fragment_analysis`, the single-fragment analysis pipeline that both
    pages call (once for SEC's one fragment, twice for RP's two fragments).
"""

from __future__ import annotations

import io
import os

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import streamlit as st

from dar_calculator import (
    MassVariant,
    PayloadDef,
    base_masses_from_mab,
    build_theoretical_table,
    build_verification_table,
    calculate_dar,
    consolidate_by_total_count,
    estimate_theoretical_grid_size,
    filter_by_fractional_abundance,
    flag_abundance_implausible,
    marginal_distribution_by_chemistry,
    match_species,
    build_drug_load_summary_table,
    build_selection_summary,
)

APP_TITLE = "DAR Compass"
APP_TAGLINE = "Automated DAR distribution analysis for antibody-drug conjugates"


# --------------------------------------------------------------------------
# Access gate
# --------------------------------------------------------------------------
# Enforced only if an APP_PASSWORD is configured (e.g. as a Render/Streamlit
# Cloud environment variable or secret). Local development with no
# APP_PASSWORD set runs ungated. Call this once, before rendering navigation,
# so it gates every page rather than needing to be repeated per page.

def _get_app_password() -> str | None:
    pw = os.environ.get("APP_PASSWORD")
    if pw:
        return pw
    try:
        return st.secrets["APP_PASSWORD"]
    except Exception:
        return None


def require_password() -> None:
    expected = _get_app_password()
    if not expected:
        return  # no password configured - nothing to gate

    if st.session_state.get("_authenticated"):
        return

    st.title(f"\U0001F9ED {APP_TITLE}")
    st.caption(APP_TAGLINE)
    st.info("This app is password-protected. Ask whoever shared the link with you for the password.")
    pw = st.text_input("Password", type="password", key="_password_input")
    if st.button("Enter", type="primary"):
        if pw == expected:
            st.session_state["_authenticated"] = True
            st.rerun()
        else:
            st.error("Incorrect password.")
    st.stop()


# --------------------------------------------------------------------------
# File loading
# --------------------------------------------------------------------------

def read_uploaded_excel(uploaded_file) -> pd.DataFrame:
    df = pd.read_excel(uploaded_file)
    required = {"Average Mass", "Sum Intensity"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"'{uploaded_file.name}' is missing expected column(s): {missing}")
    return df


# --------------------------------------------------------------------------
# Sidebar widgets shared between pages
# --------------------------------------------------------------------------

def render_chemistry_variants_ui(prefix: str, i: int) -> tuple[str, list[MassVariant]]:
    """Render the label + mass-variant widgets for one chemistry.

    Deliberately does NOT render max-conjugation-count/step - callers add
    those separately, since the SEC page needs one conjugation range per
    chemistry while the RP page needs two independent ones (one per
    fragment), sharing this same MW/variant configuration.
    """
    label = st.text_input(
        "Label", value=str(i + 1), key=f"{prefix}_label_{i}",
        help="Short tag, e.g. '4', '12', 'DXd'.",
    )
    n_variants = st.number_input(
        "Number of mass variants for this chemistry",
        min_value=1, max_value=4, value=1, step=1, key=f"{prefix}_nvariants_{i}",
        help=(
            "Usually 1 (the intact linker-payload mass). Add more if breakage during harsh sample "
            "processing produces additional possible masses at a conjugation site — DAR Compass will "
            "consider any mix of these variants across a chemistry's attachment sites within one molecule."
        ),
    )

    variants: list[MassVariant] = []
    for v in range(int(n_variants)):
        variant_tag = "intact" if v == 0 else f"variant {v + 1}"
        st.markdown(f"**Mass variant {v + 1}** ({variant_tag})" if v == 0 else f"**Mass variant {v + 1}**")
        vmw = st.number_input(
            "MW (Da)", min_value=0.0, value=1717.83 if v == 0 else 0.0, step=0.01, format="%.2f",
            key=f"{prefix}_mw_{i}_{v}",
        )
        vweight = st.number_input(
            "DAR weight", min_value=0.0, max_value=1.0, value=1.0, step=0.05, format="%.2f",
            key=f"{prefix}_weight_{i}_{v}",
            help=(
                "How much one occupied site of this mass variant counts toward DAR. 1.0 (default) = "
                "counts as a full payload, same as an intact site. Set lower (or 0) for a variant "
                "that represents partial or complete payload loss."
            ),
        )
        variants.append(MassVariant(mw=float(vmw), dar_weight=float(vweight)))
    return label, variants


def render_conjugation_range_ui(prefix: str, i: int, fragment_label: str = "") -> tuple[int, int]:
    """Render max-conjugation-count + count-step widgets for one chemistry.

    `fragment_label` (e.g. "F(ab')2" or "Fc") disambiguates the widget text
    when the same chemistry needs an independent conjugation range per
    fragment on the RP page.
    """
    suffix = f" in {fragment_label}" if fragment_label else ""
    max_n = st.number_input(
        f"Max conjugation count{suffix}", min_value=0, max_value=20, value=8, step=1,
        key=f"{prefix}_maxn_{i}",
    )
    step = st.selectbox(
        f"Allowed count step{suffix}", options=[1, 2], index=0, key=f"{prefix}_step_{i}",
        help="Use 2 if conjugation only occurs in pairs (e.g. per disulfide bond).",
    )
    return int(max_n), int(step)


# --------------------------------------------------------------------------
# Analysis orchestration (shared by both pages)
# --------------------------------------------------------------------------

def run_fragment_analysis(
    reference_file,
    sample_file,
    payload_defs: list[PayloadDef],
    adc_min_fractional_abundance: float,
    ppm_tolerance: float,
    base_mass_mode: str,
    base_mass_top_n: int,
    exclude_implausible: bool = False,
) -> dict | None:
    """Run the full matching/DAR pipeline for one fragment (or, on the SEC
    page, for the whole intact molecule - structurally identical, just one
    reference file and one sample file). Returns a results dict, or None
    (having already shown an st.error) if inputs are missing/invalid.
    """
    if reference_file is None or sample_file is None:
        st.error("Please upload both a reference (naked) file and a sample (ADC) file before running the analysis.")
        return None

    if not payload_defs:
        st.error("Please include at least one linker-payload chemistry.")
        return None

    try:
        mab_df = read_uploaded_excel(reference_file)
        adc_df_all = read_uploaded_excel(sample_file)
    except ValueError as e:
        st.error(str(e))
        return None

    try:
        adc_df = filter_by_fractional_abundance(adc_df_all, adc_min_fractional_abundance)
    except ValueError as e:
        st.error(str(e))
        return None

    if adc_df.empty:
        st.error(
            f"No peaks remain at a {adc_min_fractional_abundance:.2f}% fractional abundance threshold "
            f"(out of {len(adc_df_all)} peaks in the file). Lower the threshold and try again."
        )
        return None

    base_masses = base_masses_from_mab(mab_df, top_n=int(base_mass_top_n))
    estimated_grid_size = estimate_theoretical_grid_size(base_masses, payload_defs)
    theoretical = build_theoretical_table(base_masses, payload_defs)
    matched = match_species(adc_df, theoretical, ppm_tolerance=ppm_tolerance)
    dar_full, matched_full = calculate_dar(matched, payload_defs)
    matched_full = flag_abundance_implausible(matched_full, payload_defs)

    implausible_mask = matched_full["abundance_implausible"] if not matched_full.empty else pd.Series(dtype=bool)
    n_implausible = int(implausible_mask.sum()) if len(implausible_mask) else 0
    implausible_rows = matched_full[implausible_mask].copy() if n_implausible else matched_full.iloc[0:0].copy()

    if exclude_implausible and n_implausible:
        filtered = matched_full[~implausible_mask].copy()
        dar, matched_with_contrib = calculate_dar(filtered, payload_defs)
        matched_with_contrib = flag_abundance_implausible(matched_with_contrib, payload_defs)
    else:
        dar, matched_with_contrib = dar_full, matched_full

    n_display_cols: list[str] = []
    for p in payload_defs:
        for v in p.variants:
            n_display_cols.append(f"n_{v.variant_label}")
        if len(p.variants) > 1:
            n_display_cols.append(f"n_{p.label}_total")

    verification_df = build_verification_table(
        adc_df_all, theoretical, ppm_tolerance,
        abundance_threshold=adc_min_fractional_abundance,
    )

    return {
        "payload_defs": payload_defs,
        "payload_labels": [p.label for p in payload_defs],
        "dar": dar,
        "matched_with_contrib": matched_with_contrib,
        "n_implausible": n_implausible,
        "implausible_rows": implausible_rows,
        "exclude_implausible": exclude_implausible,
        "base_masses": base_masses,
        "base_mass_mode": base_mass_mode,
        "theoretical": theoretical,
        "verification_df": verification_df,
        "n_observed": len(adc_df_all),
        "n_candidates": len(adc_df),
        "n_matched": len(matched_with_contrib),
        "n_display_cols": n_display_cols,
        "adc_min_fractional_abundance": adc_min_fractional_abundance,
        "ppm_tolerance": ppm_tolerance,
        "estimated_grid_size": estimated_grid_size,
    }


# --------------------------------------------------------------------------
# Chart / table builders
# --------------------------------------------------------------------------

def make_distribution_chart(matched_df: pd.DataFrame, dar: dict, payload_labels: list[str], top_n: int = 15):
    fig, ax = plt.subplots(figsize=(9, 4.5))
    if matched_df.empty:
        ax.text(0.5, 0.5, "No species matched at this tolerance", ha="center", va="center")
        ax.axis("off")
        return fig

    top = matched_df.sort_values("relative_abundance", ascending=False).head(top_n)
    colors = ["#C55A11" if a else "#4472C4" for a in top["ambiguous"]]
    ax.bar(top["species"], top["relative_abundance"] * 100, color=colors)
    ax.set_ylabel("Relative abundance (%)")
    ax.set_xlabel("Assigned species")
    dar_text = "  |  ".join(f"DAR[{lbl}] = {dar[lbl]:.2f}" for lbl in payload_labels)
    ax.set_title(f"{dar_text}  |  Total DAR = {dar['total']:.2f}")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    return fig


def make_drug_load_chart(long_df: pd.DataFrame, payload_defs: list[PayloadDef], dar: dict):
    n_chem = max(len(payload_defs), 1)
    fig, axes = plt.subplots(1, n_chem, figsize=(5.5 * n_chem, 4.5))
    if n_chem == 1:
        axes = [axes]
    for ax, p in zip(axes, payload_defs):
        sub = long_df[long_df["chemistry"] == p.label].sort_values("count")
        counts = sub["count"].tolist()
        pcts = sub["relative_abundance_pct"].tolist()
        colors = ["#2A8C82"] * len(pcts)
        if pcts:
            colors[pcts.index(max(pcts))] = "#1B2A4A"  # highlight the mode, matching the reference report style
        ax.bar([str(c) for c in counts], pcts, color=colors)
        ax.set_xlabel(f"{p.label} count")
        ax.set_ylabel("Relative abundance (%)")
        ax.set_title(f"{p.label}\nAverage = {dar.get(p.label, float('nan')):.2f}")
    plt.tight_layout()
    return fig


def make_selection_funnel_chart(selection_summary: pd.DataFrame):
    """Horizontal funnel of the three stages that actually narrow the peak
    set down (skips the two "excluded"/"unmatched" bookkeeping rows, which
    are the complement of stages already shown) - lets someone see at a
    glance whether a run kept nearly all its signal or lost half of it,
    without reading percentages across a table row by row.
    """
    keep_stages = [
        "Total peaks in ADC file",
        "Candidate peaks (passed threshold)",
        "Matched to a theoretical species",
    ]
    short_labels = ["All peaks", "Above threshold", "Matched"]
    sub = selection_summary.set_index("Stage").reindex(keep_stages)
    if sub["Peak count"].isna().all():
        return None

    peak_pct = sub["% of peaks"].fillna(0).tolist()
    signal_pct = sub["% of total signal (Fractional Abundance)"].fillna(0).tolist()

    fig, ax = plt.subplots(figsize=(7, 2.2))
    y = np.arange(len(short_labels))
    bar_h = 0.32
    ax.barh(y + bar_h / 2, peak_pct, height=bar_h, color="#8FAADC", label="% of peaks")
    ax.barh(y - bar_h / 2, signal_pct, height=bar_h, color="#1B2A4A", label="% of total signal")
    for yi, (pp, sp) in enumerate(zip(peak_pct, signal_pct)):
        ax.text(pp + 1.5, yi + bar_h / 2, f"{pp:.0f}%", va="center", fontsize=9)
        ax.text(sp + 1.5, yi - bar_h / 2, f"{sp:.0f}%", va="center", fontsize=9)
    ax.set_yticks(y)
    ax.set_yticklabels(short_labels)
    ax.set_xlim(0, 115)
    ax.set_xlabel("%")
    ax.invert_yaxis()
    ax.legend(loc="lower right", fontsize=8, frameon=False)
    for spine in ("top", "right"):
        ax.spines[spine].set_visible(False)
    plt.tight_layout()
    return fig


def render_dar_sanity_badge(n_ambiguous: int, n_implausible: int, exclude_implausible: bool) -> None:
    """A small colored pill right under the DAR metrics - the one thing a
    user is guaranteed to look at - so a caveat isn't only visible to
    someone who scrolls down far enough to find the matched-species table.
    """
    if n_ambiguous == 0 and n_implausible == 0:
        bg, fg, text = "#D9EAD3", "#274E13", "Clean match — 0 ambiguous, 0 flagged"
    elif n_ambiguous == 0 and n_implausible > 0 and exclude_implausible:
        bg, fg, text = "#CFE2F3", "#1B4F72", f"0 ambiguous · {n_implausible} flagged (excluded from DAR)"
    else:
        parts = []
        if n_ambiguous:
            parts.append(f"{n_ambiguous} ambiguous")
        if n_implausible:
            parts.append(f"{n_implausible} flagged")
        bg, fg, text = "#FCE4D6", "#7F4B0D", " · ".join(parts) + " — review before trusting this DAR"

    st.markdown(
        f'<span style="background-color:{bg}; color:{fg}; padding:3px 12px; '
        f'border-radius:999px; font-size:0.85rem; font-weight:600;">{text}</span>',
        unsafe_allow_html=True,
    )


def style_drug_load_table(table: pd.DataFrame):
    count_cols = [c for c in table.columns if c != "Average"]

    def highlight_mode(row):
        styles = [""] * len(row)
        vals = row[count_cols].dropna()
        if len(vals):
            mode_col = vals.idxmax()
            styles[list(row.index).index(mode_col)] = "background-color: #CFE2F3; font-weight: bold; color: #1B2A4A;"
        return styles

    return table.style.apply(highlight_mode, axis=1).format("{:.2f}", na_rep="")


def build_params_table(
    payload_defs: list[PayloadDef],
    base_masses: list[float],
    base_mass_mode: str,
    ppm_tolerance: float,
    adc_min_fractional_abundance: float,
) -> pd.DataFrame:
    """One row per analysis parameter, in the same spirit as the header
    block of a manual analysis sheet - so someone reviewing the output can
    see exactly what was configured without having to ask.
    """
    rows = [
        {"Parameter": "mAb base mass mode", "Value": base_mass_mode},
        {"Parameter": "mAb base mass(es) used (Da)", "Value": ", ".join(f"{b:.4f}" for b in base_masses)},
        {"Parameter": "Mass accuracy tolerance (ppm)", "Value": ppm_tolerance},
        {"Parameter": "ADC fractional abundance threshold (%)", "Value": adc_min_fractional_abundance},
    ]
    for p in payload_defs:
        rows.append({"Parameter": f"Chemistry \"{p.label}\" - max conjugation count", "Value": max(p.n_values) if p.n_values else 0})
        rows.append({"Parameter": f"Chemistry \"{p.label}\" - count step", "Value": (p.n_values[1] - p.n_values[0]) if len(p.n_values) > 1 else 1})
        for v in p.variants:
            rows.append({
                "Parameter": f"Chemistry \"{p.label}\" - mass variant \"{v.variant_label}\"",
                "Value": f"MW = {v.mw:.4f} Da, DAR weight = {v.dar_weight}",
            })
    return pd.DataFrame(rows)


def write_labeled_block(ws, title: str, df: pd.DataFrame, start_row: int) -> int:
    """Write a titled table starting at `start_row`; return the next free row."""
    from openpyxl.styles import Font
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12)
    start_row += 1
    if df is None or df.empty:
        ws.cell(row=start_row, column=1, value="(none)")
        return start_row + 2

    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=j, value=str(col)).font = Font(bold=True)
    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for j, val in enumerate(row.tolist(), start=1):
            ws.cell(row=i, column=j, value=val)
    return start_row + len(df) + 3


def to_excel_bytes(
    summary_df: pd.DataFrame,
    species_df: pd.DataFrame,
    drug_load_df: pd.DataFrame | None = None,
    params_df: pd.DataFrame | None = None,
    theoretical_df: pd.DataFrame | None = None,
    verification_df: pd.DataFrame | None = None,
    selection_summary_df: pd.DataFrame | None = None,
    implausible_df: pd.DataFrame | None = None,
) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="DAR Summary", index=False)
        if selection_summary_df is not None and not selection_summary_df.empty:
            selection_summary_df.to_excel(writer, sheet_name="Selection Summary", index=False)
        if drug_load_df is not None and not drug_load_df.empty:
            drug_load_df.to_excel(writer, sheet_name="Drug-load Distribution")
        species_df.to_excel(writer, sheet_name="Matched Species", index=False)

        if params_df is not None:
            ws = writer.book.create_sheet("Analysis")
            row = write_labeled_block(ws, "Analysis Parameters", params_df, 1)
            row = write_labeled_block(ws, "Theoretical Mass Combinations (all permutations)", theoretical_df, row)
            row = write_labeled_block(
                ws,
                "Peak-by-Peak Verification (every candidate peak, matched or not)",
                verification_df,
                row,
            )
            if implausible_df is not None and not implausible_df.empty:
                write_labeled_block(
                    ws,
                    "Abundance-Implausible Species (breakage-derived, more abundant than their intact counterpart)",
                    implausible_df,
                    row,
                )
    return buf.getvalue()


# --------------------------------------------------------------------------
# Full results section (shared by both pages; called once per fragment)
# --------------------------------------------------------------------------

def render_results_section(r: dict, key_prefix: str, title_suffix: str = "") -> None:
    """Render the complete results UI - metrics, selection summary,
    drug-load distribution, species-level chart, matched species table,
    and downloads - from a `run_fragment_analysis` results dict.

    `key_prefix` keeps widget keys unique when this is called more than
    once on the same page (the RP page calls it twice: once for F(ab')2,
    once for Fc). `title_suffix` (e.g. " (F(ab')2)") is appended to labels
    for the same reason.
    """
    payload_defs = r["payload_defs"]
    payload_labels = r["payload_labels"]
    dar = r["dar"]
    matched_with_contrib = r["matched_with_contrib"]
    n_implausible = r.get("n_implausible", 0)
    implausible_rows = r.get("implausible_rows")
    exclude_implausible = r.get("exclude_implausible", False)
    base_masses = r["base_masses"]
    base_mass_mode = r["base_mass_mode"]
    theoretical = r["theoretical"]
    verification_df = r["verification_df"]
    n_observed = r["n_observed"]
    n_candidates = r["n_candidates"]
    n_matched = r["n_matched"]
    n_display_cols = r["n_display_cols"]
    adc_min_fractional_abundance = r["adc_min_fractional_abundance"]
    ppm_tolerance = r["ppm_tolerance"]
    estimated_grid_size = r["estimated_grid_size"]

    if estimated_grid_size > 300_000:
        st.warning(
            f"This configuration builds roughly {estimated_grid_size:,} theoretical species, which may run "
            "slowly. Consider reducing the number of mass variants, max conjugation count, or base masses."
        )

    with st.expander(f"Mass variants configured for this run{title_suffix}"):
        variant_rows = []
        for p in payload_defs:
            for v in p.variants:
                variant_rows.append({
                    "Chemistry": p.label, "Variant": v.variant_label,
                    "MW (Da)": v.mw, "DAR weight": v.dar_weight,
                })
        st.dataframe(
            pd.DataFrame(variant_rows), use_container_width=True, hide_index=True,
            key=f"{key_prefix}_variants_table",
        )

    with st.expander(f"Peak-by-peak verification{title_suffix} (every peak in the sample file, matched or not)"):
        st.caption(
            "Every peak from the uploaded sample file, with its closest theoretical species regardless "
            "of whether it was actually accepted - so you can check *why* a peak wasn't counted (excluded "
            "by the abundance threshold vs. simply too far in ppm from anything in the theoretical grid) "
            "instead of it just silently disappearing. Green = matched. Yellow = excluded by the "
            "fractional abundance threshold before matching was even attempted. Red = passed the "
            "abundance threshold but its closest theoretical species was still outside the ppm tolerance "
            "- if this happens for a peak you expected to match, it usually means a chemistry's max "
            "conjugation count (or a mass variant's MW) needs adjusting, not that the peak is real noise."
        )

        def highlight_verification(row):
            if row["matched"]:
                color = "#D9EAD3"  # green
            elif not row["passed_abundance_threshold"]:
                color = "#FFF2CC"  # yellow
            else:
                color = "#F4CCCC"  # red
            return [f"background-color: {color}" for _ in row]

        verification_display_cols = [
            "Average Mass", "Sum Intensity", "Fractional Abundance", "closest_species",
            "closest_theoretical_mass", "delta_mass", "ppm_error",
            "passed_abundance_threshold", "within_ppm_tolerance", "matched",
        ]
        verification_display_cols = [c for c in verification_display_cols if c in verification_df.columns]
        st.dataframe(
            verification_df[verification_display_cols].style.apply(highlight_verification, axis=1),
            use_container_width=True,
            key=f"{key_prefix}_verification_table",
        )

    # --- Top-line metrics ---------------------------------------------
    cols = st.columns(len(payload_labels) + 1)
    for col, lbl in zip(cols, payload_labels):
        col.metric(f"DAR [{lbl}]{title_suffix}", f"{dar[lbl]:.2f}")
    cols[-1].metric(f"Total DAR{title_suffix}", f"{dar['total']:.2f}")

    n_ambiguous = int(matched_with_contrib["ambiguous"].sum()) if n_matched else 0
    render_dar_sanity_badge(n_ambiguous, n_implausible, exclude_implausible)

    if n_implausible and not exclude_implausible:
        st.warning(
            f"{n_implausible} matched peak(s) flagged abundance-implausible{title_suffix}: a "
            "breakage-derived species is more abundant than the fully-intact species at the same "
            "conjugation state, which usually means the match is a spurious combinatorial "
            "coincidence rather than a real breakage species. Currently INCLUDED in the DAR "
            "calculation above - see \"Abundance-implausible species\" below, and the checkbox in "
            "the sidebar if you want to exclude them."
        )
    elif n_implausible and exclude_implausible:
        st.info(
            f"{n_implausible} peak(s) were flagged abundance-implausible{title_suffix} and have "
            "been EXCLUDED from the DAR calculation above, per the sidebar setting. See "
            "\"Abundance-implausible species\" below for what was removed."
        )

    if n_implausible:
        with st.expander(f"Abundance-implausible species{title_suffix} ({n_implausible} flagged)"):
            st.caption(
                "Breakage-derived species (using a non-intact mass variant) that are more abundant "
                "than the fully-intact species at the same total conjugation count, other "
                "chemistries held equal. Breakage is expected to be a minority pathway relative to "
                "the intact population it derives from, so this pattern usually means the match is "
                "an artifact of the denser theoretical grid multiple mass variants create, not a "
                "real species. This is a heuristic, not a hard rule - it can be wrong for a sample "
                "deliberately forced to degrade."
            )
            implausible_display_cols = [
                "Average Mass", "Sum Intensity", "species", "relative_abundance", "implausible_detail",
            ]
            implausible_display_cols = [c for c in implausible_display_cols if c in implausible_rows.columns]
            display_implausible = implausible_rows[implausible_display_cols].copy()
            if "relative_abundance" in display_implausible.columns:
                display_implausible["relative_abundance"] = (display_implausible["relative_abundance"] * 100).round(2)
                display_implausible = display_implausible.rename(columns={"relative_abundance": "relative_abundance_%"})
            st.dataframe(
                display_implausible, use_container_width=True, hide_index=True,
                key=f"{key_prefix}_implausible_table",
            )

    # --- Selection summary -----------------------------------------------
    st.subheader(f"Selection summary{title_suffix}")
    st.caption(
        "How many peaks were selected at each stage, and how much of the total signal they represent. "
        "The two percentages can diverge a lot - matching a small fraction of peaks by count can still "
        "mean capturing nearly all of the real signal, since low-abundance noise peaks are expected to "
        "go unmatched. See \"Peak-by-peak verification\" above for the detail behind these numbers."
    )
    selection_summary = build_selection_summary(verification_df)
    funnel_fig = make_selection_funnel_chart(selection_summary) if not selection_summary.empty else None
    if funnel_fig is not None:
        st.pyplot(funnel_fig)
    st.dataframe(
        selection_summary, use_container_width=True, hide_index=True,
        key=f"{key_prefix}_selection_summary",
    )

    if n_ambiguous:
        st.warning(
            f"{n_ambiguous} matched peak(s) had more than one plausible species within tolerance. "
            "Review the highlighted rows below before trusting the DAR value."
        )

    # --- Drug-load distribution (per-chemistry summary table + chart) ----
    st.subheader(f"Drug-load distribution{title_suffix}")
    st.caption(
        "Percent of matched intensity at each individual count, per chemistry - independent of the "
        "other chemistries' counts. This is the standard drug-load-distribution report format (one "
        "row per linker-payload, one column per count, plus an average)."
    )

    drug_load_long = marginal_distribution_by_chemistry(matched_with_contrib, payload_defs)
    drug_load_table = build_drug_load_summary_table(matched_with_contrib, payload_defs, dar)

    has_nondefault_dar_weight = any(v.dar_weight != 1.0 for p in payload_defs for v in p.variants)
    if has_nondefault_dar_weight:
        st.caption(
            "Note: at least one mass variant has a DAR weight other than 1.0. \"Average\" is each "
            "chemistry's actual DAR (weighted by dar_weight, same number as the metric above) - it will "
            "differ from the simple mean of the row above it, which counts every occupied site equally "
            "regardless of variant."
        )

    if not drug_load_table.empty:
        st.dataframe(style_drug_load_table(drug_load_table), use_container_width=True, key=f"{key_prefix}_drug_load_table")
        drug_load_fig = make_drug_load_chart(drug_load_long, payload_defs, dar)
        st.pyplot(drug_load_fig)
    else:
        drug_load_fig = None
        st.write("No matched species to summarize yet.")

    # --- Species-level chart ----------------------------------------------
    st.subheader(f"DAR distribution (species-level){title_suffix}")

    has_multi_variant_chemistry = any(len(p.variants) > 1 for p in payload_defs)
    chart_df = matched_with_contrib
    if has_multi_variant_chemistry:
        chart_view = st.radio(
            "Chart detail",
            options=["Detailed (show mass variants)", "Consolidated (total counts per chemistry)"],
            horizontal=True,
            key=f"{key_prefix}_chart_view",
            help=(
                "Detailed shows every intact/broken mixture as its own bar - precise, but can get "
                "cluttered once a chemistry has multiple mass variants. Consolidated groups bars by "
                "total occupied-site count per chemistry only, hiding which specific variant(s) made "
                "up that total. This toggle only changes this chart - the table below and the DAR "
                "numbers above are unaffected either way."
            ),
        )
        if chart_view.startswith("Consolidated"):
            chart_df = consolidate_by_total_count(matched_with_contrib, payload_defs)

    fig = make_distribution_chart(chart_df, dar, payload_labels)
    st.pyplot(fig)

    # --- Matched species table --------------------------------------------
    st.subheader(f"Matched species{title_suffix}")
    if n_matched:
        display_cols = ["Average Mass", "Sum Intensity", "species", "ppm_error", "relative_abundance"]
        display_cols += n_display_cols
        display_cols += ["ambiguous", "runner_up_species", "runner_up_ppm_error"]
        if "abundance_implausible" in matched_with_contrib.columns:
            display_cols += ["abundance_implausible"]
        display_df = matched_with_contrib[display_cols].copy()
        display_df["relative_abundance"] = (display_df["relative_abundance"] * 100).round(2)
        display_df = display_df.rename(columns={"relative_abundance": "relative_abundance_%"})

        def highlight_row(row):
            if row.get("abundance_implausible"):
                color = "#E6D9F2"  # purple - abundance-implausible takes visual priority
            elif row["ambiguous"]:
                color = "#FCE4D6"  # orange - mass-ambiguous
            else:
                color = ""
            return [f"background-color: {color}" if color else "" for _ in row]

        st.dataframe(
            display_df.style.apply(highlight_row, axis=1), use_container_width=True,
            key=f"{key_prefix}_matched_species_table",
        )
    else:
        st.write("No peaks matched at this tolerance. Try increasing the ppm tolerance.")

    # --- Downloads -----------------------------------------------------------
    st.subheader(f"Download results{title_suffix}")
    summary_df = pd.DataFrame(
        {
            "metric": [f"DAR [{lbl}]" for lbl in payload_labels] + [
                "Total DAR", "Matched peaks", "Candidate peaks (after abundance filter)",
                "Observed peaks (raw file)", "ADC fractional abundance threshold (%)",
                "Ambiguous matches", "Abundance-implausible flagged", "Abundance-implausible excluded from DAR",
                "ppm tolerance used",
            ],
            "value": [round(dar[lbl], 3) for lbl in payload_labels] + [
                round(dar["total"], 3), n_matched, n_candidates,
                n_observed, adc_min_fractional_abundance,
                n_ambiguous, n_implausible, exclude_implausible,
                ppm_tolerance,
            ],
        }
    )

    params_df = build_params_table(payload_defs, base_masses, base_mass_mode, ppm_tolerance, adc_min_fractional_abundance)
    file_tag = key_prefix.replace(" ", "_")

    col1, col2, col3 = st.columns(3)
    with col1:
        st.download_button(
            "Download DAR report (.xlsx)",
            data=to_excel_bytes(
                summary_df, matched_with_contrib, drug_load_table,
                params_df=params_df, theoretical_df=theoretical, verification_df=verification_df,
                selection_summary_df=selection_summary, implausible_df=implausible_rows,
            ),
            file_name=f"dar_report_{file_tag}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key=f"{key_prefix}_download_xlsx",
            help=(
                "Includes the drug-load distribution table and an \"Analysis\" sheet with every "
                "parameter used, the full theoretical mass grid, and a peak-by-peak verification "
                "table - so results can be checked by hand."
            ),
        )
    with col2:
        img_buf = io.BytesIO()
        fig.savefig(img_buf, format="png", dpi=150)
        st.download_button(
            "Download species-level chart (.png)",
            data=img_buf.getvalue(),
            file_name=f"dar_distribution_species_{file_tag}.png",
            mime="image/png",
            use_container_width=True,
            key=f"{key_prefix}_download_species_png",
        )
    with col3:
        if drug_load_fig is not None:
            drug_load_img_buf = io.BytesIO()
            drug_load_fig.savefig(drug_load_img_buf, format="png", dpi=150)
            st.download_button(
                "Download drug-load chart (.png)",
                data=drug_load_img_buf.getvalue(),
                file_name=f"drug_load_distribution_{file_tag}.png",
                mime="image/png",
                use_container_width=True,
                key=f"{key_prefix}_download_drugload_png",
            )

    with st.expander(f"Base masses used for this run{title_suffix}"):
        st.write(pd.DataFrame({"base_mass": base_masses}))
